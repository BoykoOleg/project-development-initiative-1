"""Чтение и преобразование Excel (.xlsx) файлов в текст для ИИ"""
import io
import requests

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

MAX_ROWS = 200      # максимум строк на лист
MAX_COLS = 30       # максимум колонок
MAX_SHEETS = 5      # максимум листов
MAX_CHARS = 12000   # максимум символов итогового текста


def download_file(bot_token: str, file_id: str) -> bytes:
    """Скачивает файл из Telegram по file_id"""
    file_info = requests.get(
        f"https://api.telegram.org/bot{bot_token}/getFile",
        params={"file_id": file_id},
        timeout=15
    ).json()
    file_path = file_info["result"]["file_path"]
    file_url = f"https://api.telegram.org/file/bot{bot_token}/{file_path}"
    return requests.get(file_url, timeout=30).content


def _cell_value(cell) -> str:
    """Преобразует значение ячейки в строку"""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def xlsx_to_text(file_bytes: bytes, filename: str = "файл.xlsx") -> str:
    """
    Читает xlsx и возвращает текстовое представление всех листов.
    Оптимизировано для передачи в контекст ИИ.
    """
    if not OPENPYXL_OK:
        return "[Ошибка: библиотека openpyxl не установлена]"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return f"[Ошибка чтения файла: {e}]"

    parts = [f"📊 Файл: {filename}"]
    sheets_processed = 0

    for sheet_name in wb.sheetnames[:MAX_SHEETS]:
        ws = wb[sheet_name]
        rows_data = []
        row_count = 0

        for row in ws.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS):
            cells = [_cell_value(c) for c in row]
            # Пропускаем полностью пустые строки
            if not any(cells):
                continue
            rows_data.append(cells)
            row_count += 1
            if row_count >= MAX_ROWS:
                break

        if not rows_data:
            continue

        sheets_processed += 1
        parts.append(f"\n--- Лист: {sheet_name} ({row_count} строк) ---")

        # Первая строка — заголовки (если непустые)
        headers = rows_data[0] if rows_data else []
        has_headers = any(headers)

        for i, row_cells in enumerate(rows_data):
            # Убираем trailing пустые ячейки
            while row_cells and not row_cells[-1]:
                row_cells.pop()
            if not row_cells:
                continue

            if has_headers and i == 0:
                parts.append("Заголовки: " + " | ".join(row_cells))
            else:
                if has_headers and len(headers) == len(row_cells):
                    # Форматируем как ключ=значение
                    pairs = [f"{headers[j]}={row_cells[j]}" for j in range(len(row_cells)) if row_cells[j]]
                    parts.append(", ".join(pairs))
                else:
                    parts.append(" | ".join(row_cells))

    wb.close()

    if sheets_processed == 0:
        return f"[Файл {filename} пустой или не содержит данных]"

    result = "\n".join(parts)
    if len(result) > MAX_CHARS:
        result = result[:MAX_CHARS] + f"\n[... данные обрезаны, показано {MAX_CHARS} символов из {len(result)}]"

    return result


def process_excel_document(bot_token: str, message: dict) -> str | None:
    """
    Проверяет, является ли документ xlsx-файлом, и возвращает его текстовое содержимое.
    Возвращает None если документ не является Excel.
    """
    document = message.get("document")
    if not document:
        return None

    mime = document.get("mime_type", "")
    filename = document.get("file_name", "файл")

    is_xlsx = (
        mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or filename.lower().endswith(".xlsx")
        or filename.lower().endswith(".xls")
    )

    if not is_xlsx:
        return None

    file_id = document["file_id"]
    file_bytes = download_file(bot_token, file_id)
    return xlsx_to_text(file_bytes, filename)
